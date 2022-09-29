from importlib.metadata import files
from msilib.schema import File
from openpyxl import load_workbook
from googletrans import Translator

lastrow = 50         # last row of sheet
owordcol = 2          # original word column
twordcol = 3          # translated word column
owordlang = 'zh-tw'   # original word language
twordlang = 'ko'      # translated word language
file = 'test.xlsx'   # target excel file
sheetno = 'test'     # target excel sheet


wb = load_workbook(file) 
ws = wb[sheetno]                 
ts = Translator()

for i in range (1, lastrow, 1):      
    oword = str(ws.cell(i,owordcol).value)
    print(oword)
    tword = ts.translate(oword, src=owordlang, dest=twordlang).text
    ws.cell(i,twordcol).value = tword
wb.save(file)          